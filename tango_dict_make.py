import openpyxl

wb = openpyxl.load_workbook("tango.xlsx")
ws = wb["Sheet1"]

tango_dict = dict()
def make():
    for row in ws.rows:
        all_data = []
        for cell in row:
            all_data.append(cell.coordinate)
        clm_123 = []
        for i in range(1,4):
            print(str(ws[all_data[i]].value))
            clm_123.append(str(ws[all_data[i]].value))
        print(clm_123)
        tango_dict[str(ws[all_data[0]].value)] = clm_123
    return tango_dict
    
